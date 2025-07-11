import openpyxl

# 1. 打开 xlsx 文件。
wb = openpyxl.load_workbook('test.xlsx')

# 2. 获取到指定的标签页
sheet = wb.active  # 默认获取第一个工作表，也可以通过名字获取：wb['SheetName']

# 3. 获取到表格中有多少行
nrows = sheet.max_row

# 4. 进行循环统计操作
total = 0
count = 0

for i in range(2, nrows + 1):  # 假设第一行为标题行
    # 拿到当前同学的班级 id
    classId = sheet.cell(row=i, column=2).value

    if classId == 100:
        total += sheet.cell(row=i, column=3).value
        count += 1

if count > 0:
    print(f'平均分：{total / count}')
else:
    print('没有找到符合条件的数据')



